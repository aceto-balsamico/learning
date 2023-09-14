
# import openpyxl

# def readData(filename):
#     with open(filename, 'r') as file:
#         hexData = file.read().split()

#     return [int(x, 16) for x in hexData]

# def drawToExcel(data):
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Original Data"

#     row, col = 2, 2  # B2セルから始める
#     for byte in data:
#         for bit in range(7, -1, -1):
#             value = (byte >> bit) & 1
#             cell = sheet.cell(row=row, column=col)
#             if value == 1:
#                 cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # 赤
#             else:
#                 cell.fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid") # 黒

#             col += 1

#             if col > 17:  # 16Byte + 数値列
#                 col = 2  # リセット
#                 row += 1

#     # 新しいシートを作成
#     new_sheet = workbook.create_sheet("Interpreted Data")

#     # データを解釈して新しいシートに描画
#     row, col = 2, 2
#     for byte in data:
#         if byte != 0xFF:
#             cell_value = format(byte, '02X')  # 16進数文字列に変換
#             cell = new_sheet.cell(row=row, column=col, value=cell_value)
#             cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # 赤
#         else:
#             cell_value = 'FF'
#             cell = new_sheet.cell(row=row, column=col, value=cell_value)
#             cell.fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid") # 黒

#         col += 1

#         if col > 17:
#             col = 2
#             row += 1

#     # シート1に数値を追加
#     for i in range(len(data)//16):
#         cell = sheet.cell(row=i+2, column=1, value=format(i*16, '02X'))
#     for i in range(len(data)//16):
#         cell = sheet.cell(row=1, column=i+2, value=i) 

#     # シート2に数値を追加
#     for i in range(len(data)//16):
#         cell = new_sheet.cell(row=i+2, column=1, value=i*16) 
#     for i in range(len(data)//16):
#         cell = new_sheet.cell(row=1, column=i+2, value=i) 
        
# 	# 	    # 表示範囲を拡張
#     # sheet.calculate_dimension()
#     # new_sheet.calculate_dimension()

#     workbook.save('output.xlsx')

# data = readData('data.txt')
# drawToExcel(data)
import openpyxl

def readData(filename):
    with open(filename, 'r') as file:
        hexData = file.read().split()

    return [int(x, 16) for x in hexData]

def drawToExcel(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Original Data"

    row, col = 2, 2  # B2セルから始める
    for byte in data:
        for bit in range(7, -1, -1):
            value = (byte >> bit) & 1
            cell = sheet.cell(row=row, column=col)
            if value == 1:
                cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # 赤
            else:
                cell.fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid") # 黒

            col += 1

            if col > 9:  # 8bit + 数値列
                col = 2  # リセット
                row += 1

    # 新しいシートを作成
    new_sheet = workbook.create_sheet("Interpreted Data")

    # データを解釈して新しいシートに描画
    row, col = 2, 2
    for byte in data:
        if byte != 0xFF:
            cell_value = format(byte, '02X')  # 16進数文字列に変換
            cell = new_sheet.cell(row=row, column=col, value=cell_value)
            cell.fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # 赤
        else:
            cell_value = 'FF'
            cell = new_sheet.cell(row=row, column=col, value=cell_value)
            cell.fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid") # 黒

        col += 1

        if col > 17:
            col = 2
            row += 1

    # シート1に数値を追加
    for i in range(len(data)):
        cell = sheet.cell(row=i+2, column=1, value='0x'+format(i*16, '02X'))
    for i in range(8):
        cell = sheet.cell(row=1, column=i+2, value='0x'+format(i, '02X')) 

    # シート2に数値を追加
    for i in range(len(data)//16):
        cell = new_sheet.cell(row=i+2, column=1, value='0x'+format(i*16, '04X'))
    for i in range(16):
        cell = new_sheet.cell(row=1, column=i+2, value='0x'+format(i, '02X'))

    # 表示範囲を拡張
    sheet.calculate_dimension()
    new_sheet.calculate_dimension()

    workbook.save('output.xlsx')

data = readData('data.txt')
drawToExcel(data)
